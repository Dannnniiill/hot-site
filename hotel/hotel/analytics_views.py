import os
from collections import defaultdict
from datetime import datetime, timedelta, time
from io import BytesIO
from pathlib import Path

import reportlab
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import simpleSplit
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import BookingDate, PromotionEvent


def parse_date_safe(value):
    if not value:
        return None

    value = str(value).strip()

    for fmt in ['%Y-%m-%d', '%d.%m.%Y']:
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            continue

    return None


def make_datetime_range(date_from, date_to):
    start_dt = None
    end_dt = None

    if date_from:
        start_dt = datetime.combine(date_from, time.min)
        if timezone.is_naive(start_dt):
            start_dt = timezone.make_aware(start_dt, timezone.get_current_timezone())

    if date_to:
        end_dt = datetime.combine(date_to, time.max)
        if timezone.is_naive(end_dt):
            end_dt = timezone.make_aware(end_dt, timezone.get_current_timezone())

    return start_dt, end_dt


def format_period_label(date_value, group_by):
    if group_by == 'month':
        return date_value.strftime('%m.%Y')
    if group_by == 'year':
        return date_value.strftime('%Y')
    return date_value.strftime('%d.%m.%Y')


def get_period_key(date_value, group_by):
    if group_by == 'month':
        return date_value.replace(day=1)
    if group_by == 'year':
        return date_value.replace(month=1, day=1)
    return date_value


def normalize_status(status_value):
    status_value = str(status_value or '').strip().lower()

    if status_value in ['cancelled', 'canceled', 'cancel', 'отменена']:
        return 'cancelled'

    if status_value in ['confirmed', 'approved', 'accepted', 'подтверждена']:
        return 'confirmed'

    return 'new'


def get_booking_reference_date(booking):
    if booking.start_date:
        return booking.start_date

    if booking.created_at:
        return timezone.localtime(booking.created_at).date()

    return None


def filter_bookings_by_period(bookings_qs, date_from=None, date_to=None):
    if date_from:
        bookings_qs = bookings_qs.filter(
            Q(start_date__isnull=False, start_date__gte=date_from) |
            Q(start_date__isnull=True, created_at__date__gte=date_from)
        )

    if date_to:
        bookings_qs = bookings_qs.filter(
            Q(start_date__isnull=False, start_date__lte=date_to) |
            Q(start_date__isnull=True, created_at__date__lte=date_to)
        )

    return bookings_qs


def clean_text(value):
    if value is None:
        return ''

    value = str(value).strip()

    if value.lower() in ['', 'undefined', 'null', 'none']:
        return ''

    return value


def build_client_name(first_name, last_name, email=''):
    first_name = clean_text(first_name)
    last_name = clean_text(last_name)
    email = clean_text(email)

    full_name = f'{first_name} {last_name}'.strip()

    if full_name:
        return full_name

    if email and '@' in email:
        return email.split('@')[0]

    if email:
        return email

    return 'Без имени'


def get_analytics_data(date_from=None, date_to=None, group_by='day'):
    if group_by not in ['day', 'month', 'year']:
        group_by = 'day'

    if not date_to:
        date_to = timezone.localdate()

    if not date_from:
        date_from = date_to - timedelta(days=30)

    start_dt, end_dt = make_datetime_range(date_from, date_to)

    bookings_all = BookingDate.objects.all()
    promo_events_all = PromotionEvent.objects.all()

    bookings_qs = filter_bookings_by_period(bookings_all, date_from, date_to)
    promo_events_qs = promo_events_all

    if start_dt:
        promo_events_qs = promo_events_qs.filter(created_at__gte=start_dt)

    if end_dt:
        promo_events_qs = promo_events_qs.filter(created_at__lte=end_dt)

    total_bookings = bookings_qs.count()
    active_bookings = bookings_qs.exclude(
        status__in=['cancelled', 'canceled', 'cancel', 'отменена']
    ).count()
    canceled_bookings = bookings_qs.filter(
        status__in=['cancelled', 'canceled', 'cancel', 'отменена']
    ).count()

    total_revenue = bookings_qs.exclude(
        status__in=['cancelled', 'canceled', 'cancel', 'отменена']
    ).aggregate(total=Sum('total_price'))['total'] or 0

    average_booking_amount = round(total_revenue / active_bookings, 2) if active_bookings > 0 else 0

    promo_bookings_count = bookings_qs.exclude(promo_code__isnull=True).exclude(promo_code='').count()

    bookings_chart_map = defaultdict(lambda: {'bookings': 0, 'canceled': 0, 'revenue': 0})
    for booking in bookings_qs:
        booking_date = get_booking_reference_date(booking)
        if not booking_date:
            continue

        period_key = get_period_key(booking_date, group_by)
        bookings_chart_map[period_key]['bookings'] += 1

        if normalize_status(booking.status) == 'cancelled':
            bookings_chart_map[period_key]['canceled'] += 1
        else:
            bookings_chart_map[period_key]['revenue'] += int(booking.total_price or 0)

    bookings_chart = [
        {
            'label': format_period_label(period_key, group_by),
            'bookings': bookings_chart_map[period_key]['bookings'],
            'canceled': bookings_chart_map[period_key]['canceled'],
            'revenue': bookings_chart_map[period_key]['revenue'],
        }
        for period_key in sorted(bookings_chart_map.keys())
    ]

    room_type_distribution_raw = (
        bookings_qs.values('type')
        .annotate(value=Count('id'))
        .order_by('-value')
    )
    room_type_distribution = [
        {
            'name': clean_text(item['type']) or 'Не указано',
            'value': item['value'],
        }
        for item in room_type_distribution_raw
    ]

    promotions_report_raw = (
        bookings_qs.exclude(promo_code__isnull=True)
        .exclude(promo_code='')
        .values('promo_code')
        .annotate(
            usages=Count('id'),
            total_discount=Sum('promo_discount'),
            total_revenue=Sum('total_price'),
        )
        .order_by('-usages', '-total_revenue')
    )
    promotions_report = [
        {
            'promo_code': clean_text(item['promo_code']),
            'usages': item['usages'] or 0,
            'total_discount': item['total_discount'] or 0,
            'total_revenue': item['total_revenue'] or 0,
        }
        for item in promotions_report_raw
    ]

    top_clients_raw = (
        bookings_qs.values('email', 'first_name', 'last_name')
        .annotate(
            bookings_count=Count('id'),
            cancellations=Count(
                'id',
                filter=Q(status__in=['cancelled', 'canceled', 'cancel', 'отменена'])
            ),
            total_spent=Sum(
                'total_price',
                filter=~Q(status__in=['cancelled', 'canceled', 'cancel', 'отменена'])
            ),
        )
        .order_by('-bookings_count', '-total_spent')[:10]
    )
    top_clients = [
        {
            'client': build_client_name(item['first_name'], item['last_name'], item['email']),
            'email': clean_text(item['email']),
            'bookings_count': item['bookings_count'] or 0,
            'cancellations': item['cancellations'] or 0,
            'average_check': round((item['total_spent'] or 0) / (item['bookings_count'] or 1), 2),
            'total_spent': item['total_spent'] or 0,
        }
        for item in top_clients_raw
    ]

    bookings_table = []
    for booking in bookings_qs.order_by('-created_at')[:20]:
        bookings_table.append({
            'booking_number': booking.booking_number or f'HTL-{booking.id:06d}',
            'client': build_client_name(booking.first_name, booking.last_name, booking.email),
            'email': clean_text(booking.email),
            'type': clean_text(booking.type),
            'amount': booking.amount or 0,
            'nights': booking.nights or 0,
            'total_price': int(booking.total_price or 0),
            'status': clean_text(booking.status),
            'start_date': str(booking.start_date or ''),
            'end_date': str(booking.end_date or ''),
            'created_at': booking.created_at.isoformat() if booking.created_at else '',
        })

    promotion_events_chart_map = defaultdict(lambda: {'apply': 0, 'copy': 0, 'details': 0})
    for event in promo_events_qs:
        event_date = timezone.localtime(event.created_at).date() if event.created_at else None
        if not event_date:
            continue

        period_key = get_period_key(event_date, group_by)
        event_type = str(event.event_type or '').strip().lower()

        if event_type in ['apply', 'copy', 'details']:
            promotion_events_chart_map[period_key][event_type] += 1

    promotion_events_chart = [
        {
            'label': format_period_label(period_key, group_by),
            'apply': promotion_events_chart_map[period_key]['apply'],
            'copy': promotion_events_chart_map[period_key]['copy'],
            'details': promotion_events_chart_map[period_key]['details'],
        }
        for period_key in sorted(promotion_events_chart_map.keys())
    ]

    promotion_events_summary = {
        'apply_count': promo_events_qs.filter(event_type='apply').count(),
        'copy_count': promo_events_qs.filter(event_type='copy').count(),
        'details_count': promo_events_qs.filter(event_type='details').count(),
        'total_events': promo_events_qs.count(),
    }

    promotion_events_table_raw = (
        promo_events_qs.values('promo_code', 'promo_title', 'discount_label')
        .annotate(
            apply_count=Count('id', filter=Q(event_type='apply')),
            copy_count=Count('id', filter=Q(event_type='copy')),
            details_count=Count('id', filter=Q(event_type='details')),
            total_events=Count('id'),
        )
        .order_by('-total_events', 'promo_code')
    )
    promotion_events_table = [
        {
            'promo_code': clean_text(item['promo_code']),
            'promo_title': clean_text(item['promo_title']),
            'discount_label': clean_text(item['discount_label']),
            'apply_count': item['apply_count'] or 0,
            'copy_count': item['copy_count'] or 0,
            'details_count': item['details_count'] or 0,
            'total_events': item['total_events'] or 0,
        }
        for item in promotion_events_table_raw
    ]

    return {
        'debug': {
            'bookings_all_count': bookings_all.count(),
            'bookings_filtered_count': bookings_qs.count(),
            'promo_events_all_count': promo_events_all.count(),
            'promo_events_filtered_count': promo_events_qs.count(),
            'start_dt': start_dt.isoformat() if start_dt else None,
            'end_dt': end_dt.isoformat() if end_dt else None,
        },
        'period': {
            'from': str(date_from),
            'to': str(date_to),
            'group_by': group_by,
        },
        'summary': {
            'total_bookings': total_bookings,
            'active_bookings': active_bookings,
            'canceled_bookings': canceled_bookings,
            'total_revenue': total_revenue,
            'average_booking_amount': average_booking_amount,
            'promo_bookings_count': promo_bookings_count,
        },
        'charts': {
            'bookings_chart': bookings_chart,
            'room_type_distribution': room_type_distribution,
            'promotion_events_chart': promotion_events_chart,
        },
        'tables': {
            'promotions_report': promotions_report,
            'top_clients': top_clients,
            'bookings_table': bookings_table,
            'promotion_events_table': promotion_events_table,
        },
        'promotion_events_summary': promotion_events_summary,
    }


def register_pdf_fonts():
    current_dir = Path(__file__).resolve().parent
    project_fonts_dir = current_dir / 'fonts'

    font_pairs = []

    project_arial = project_fonts_dir / 'arial.ttf'
    project_arial_bold = project_fonts_dir / 'arialbd.ttf'

    if project_arial.exists() and project_arial_bold.exists():
        font_pairs.append(
            (
                str(project_arial),
                str(project_arial_bold),
                'ProjectArial',
                'ProjectArial-Bold',
            )
        )
    elif project_arial.exists():
        font_pairs.append(
            (
                str(project_arial),
                str(project_arial),
                'ProjectArial',
                'ProjectArial-Bold',
            )
        )

    font_pairs.extend([
        (
            'C:/Windows/Fonts/arial.ttf',
            'C:/Windows/Fonts/arialbd.ttf',
            'ArialUnicode',
            'ArialUnicode-Bold',
        ),
        (
            'C:/Windows/Fonts/calibri.ttf',
            'C:/Windows/Fonts/calibrib.ttf',
            'CalibriUnicode',
            'CalibriUnicode-Bold',
        ),
        (
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
            'DejaVuSans',
            'DejaVuSans-Bold',
        ),
    ])

    reportlab_fonts_dir = Path(reportlab.__file__).resolve().parent / 'fonts'
    vera_regular = reportlab_fonts_dir / 'Vera.ttf'
    vera_bold = reportlab_fonts_dir / 'VeraBd.ttf'

    if vera_regular.exists() and vera_bold.exists():
        font_pairs.append(
            (
                str(vera_regular),
                str(vera_bold),
                'Vera',
                'Vera-Bold',
            )
        )

    for regular_path, bold_path, regular_name, bold_name in font_pairs:
        if os.path.exists(regular_path) and os.path.exists(bold_path):
            try:
                if regular_name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(regular_name, regular_path))
                if bold_name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                return regular_name, bold_name
            except Exception:
                continue

    return 'Helvetica', 'Helvetica-Bold'


class AnalyticsDashboardView(APIView):
    def get(self, request):
        date_from = parse_date_safe(request.GET.get('date_from'))
        date_to = parse_date_safe(request.GET.get('date_to'))
        group_by = str(request.GET.get('group_by', 'day')).strip().lower()

        data = get_analytics_data(
            date_from=date_from,
            date_to=date_to,
            group_by=group_by,
        )
        return Response(data)


class AnalyticsBookingsView(APIView):
    def get(self, request):
        date_from = parse_date_safe(request.GET.get('date_from'))
        date_to = parse_date_safe(request.GET.get('date_to'))

        bookings_qs = filter_bookings_by_period(BookingDate.objects.all(), date_from, date_to)

        data = []
        for booking in bookings_qs.order_by('-created_at'):
            data.append({
                'id': booking.id,
                'booking_number': booking.booking_number or f'HTL-{booking.id:06d}',
                'first_name': clean_text(booking.first_name),
                'last_name': clean_text(booking.last_name),
                'email': clean_text(booking.email),
                'phone_number': clean_text(booking.phone_number),
                'type': clean_text(booking.type),
                'number': clean_text(booking.number),
                'amount': booking.amount or 0,
                'nights': booking.nights or 0,
                'promo_code': clean_text(booking.promo_code),
                'promo_discount': int(booking.promo_discount or 0),
                'total_price': int(booking.total_price or 0),
                'status': clean_text(booking.status),
                'start_date': str(booking.start_date or ''),
                'end_date': str(booking.end_date or ''),
                'created_at': booking.created_at.isoformat() if booking.created_at else '',
            })

        return Response(data)


class AnalyticsExportExcelView(APIView):
    def get(self, request):
        date_from = parse_date_safe(request.GET.get('date_from'))
        date_to = parse_date_safe(request.GET.get('date_to'))
        group_by = str(request.GET.get('group_by', 'day')).strip().lower()

        data = get_analytics_data(
            date_from=date_from,
            date_to=date_to,
            group_by=group_by,
        )

        wb = Workbook()

        header_fill = PatternFill(fill_type='solid', fgColor='8B5E3C')
        header_font = Font(color='FFFFFF', bold=True)
        title_font = Font(bold=True, size=14)

        ws = wb.active
        ws.title = 'Общая аналитика'

        ws['A1'] = 'Отчёт по аналитике сайта отеля'
        ws['A1'].font = title_font

        ws['A3'] = 'Период'
        ws['B3'] = f"{data['period']['from']} - {data['period']['to']}"

        ws['A5'] = 'Показатель'
        ws['B5'] = 'Значение'

        summary_rows = [
            ('Всего бронирований', data['summary']['total_bookings']),
            ('Активные', data['summary']['active_bookings']),
            ('Отменённые', data['summary']['canceled_bookings']),
            ('Выручка', data['summary']['total_revenue']),
            ('Средний чек', data['summary']['average_booking_amount']),
            ('С промокодом', data['summary']['promo_bookings_count']),
        ]

        row_num = 6
        for label, value in summary_rows:
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'] = value
            row_num += 1

        for cell in ws[5]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

        ws2 = wb.create_sheet('Популярность номеров')
        ws2.append(['Тип номера', 'Количество бронирований'])
        for row in data['charts']['room_type_distribution']:
            ws2.append([row['name'], row['value']])

        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 25

        ws3 = wb.create_sheet('Промокоды')
        ws3.append(['Промокод', 'Использований', 'Сумма скидок', 'Доход'])
        for row in data['tables']['promotions_report']:
            ws3.append([
                row['promo_code'],
                row['usages'],
                row['total_discount'],
                row['total_revenue'],
            ])

        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 18
        ws3.column_dimensions['D'].width = 18

        ws4 = wb.create_sheet('Топ клиентов')
        ws4.append(['Клиент', 'Email', 'Бронирований', 'Отмен', 'Средний чек', 'Потрачено'])
        for row in data['tables']['top_clients']:
            ws4.append([
                row['client'],
                row['email'],
                row['bookings_count'],
                row['cancellations'],
                row['average_check'],
                row['total_spent'],
            ])

        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws4.column_dimensions['A'].width = 28
        ws4.column_dimensions['B'].width = 32
        ws4.column_dimensions['C'].width = 16
        ws4.column_dimensions['D'].width = 14
        ws4.column_dimensions['E'].width = 18
        ws4.column_dimensions['F'].width = 18

        ws5 = wb.create_sheet('Последние бронирования')
        ws5.append(['№ брони', 'Клиент', 'Email', 'Тип', 'Сумма', 'Статус', 'Заезд', 'Выезд'])
        for row in data['tables']['bookings_table']:
            ws5.append([
                row['booking_number'],
                row['client'],
                row['email'],
                row['type'],
                row['total_price'],
                row['status'],
                row['start_date'],
                row['end_date'],
            ])

        for cell in ws5[1]:
            cell.fill = header_fill
            cell.font = header_font

        ws5.column_dimensions['A'].width = 18
        ws5.column_dimensions['B'].width = 26
        ws5.column_dimensions['C'].width = 30
        ws5.column_dimensions['D'].width = 16
        ws5.column_dimensions['E'].width = 14
        ws5.column_dimensions['F'].width = 14
        ws5.column_dimensions['G'].width = 14
        ws5.column_dimensions['H'].width = 14

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"analytics_report_{data['period']['from']}_{data['period']['to']}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class AnalyticsExportPdfView(APIView):
    def get(self, request):
        date_from = parse_date_safe(request.GET.get('date_from'))
        date_to = parse_date_safe(request.GET.get('date_to'))
        group_by = str(request.GET.get('group_by', 'day')).strip().lower()

        data = get_analytics_data(
            date_from=date_from,
            date_to=date_to,
            group_by=group_by,
        )

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        font_regular, font_bold = register_pdf_fonts()

        left_x = 20 * mm
        right_x = width - 20 * mm
        usable_width = right_x - left_x
        y = height - 20 * mm

        def new_page():
            nonlocal y
            p.showPage()
            y = height - 20 * mm
            p.setFont(font_regular, 11)

        def ensure_space(lines_count=1, line_height=6 * mm, extra=0):
            nonlocal y
            needed = lines_count * line_height + extra
            if y - needed < 20 * mm:
                new_page()

        def draw_text(text, bold=False, size=11, line_height=6 * mm, gap_after=0):
            nonlocal y
            text = str(text or '')
            font_name = font_bold if bold else font_regular
            lines = simpleSplit(text, font_name, size, usable_width)
            ensure_space(len(lines), line_height, gap_after)

            p.setFont(font_name, size)
            for line in lines:
                p.drawString(left_x, y, line)
                y -= line_height

            if gap_after:
                y -= gap_after

        def draw_section_title(text):
            draw_text(text, bold=True, size=13, line_height=7 * mm, gap_after=1 * mm)

        def money(value):
            try:
                value = float(value or 0)
                if value.is_integer():
                    return f"{int(value)} руб."
                return f"{round(value, 2)} руб."
            except Exception:
                return "0 руб."

        draw_text('Отчёт по аналитике сайта отеля', bold=True, size=16, line_height=8 * mm, gap_after=2 * mm)
        draw_text(f"Период: {data['period']['from']} - {data['period']['to']}", size=11, gap_after=3 * mm)

        draw_section_title('Общая сводка')
        summary_lines = [
            f"Всего бронирований: {data['summary']['total_bookings']}",
            f"Активные: {data['summary']['active_bookings']}",
            f"Отменённые: {data['summary']['canceled_bookings']}",
            f"Выручка: {money(data['summary']['total_revenue'])}",
            f"Средний чек: {money(data['summary']['average_booking_amount'])}",
            f"С промокодом: {data['summary']['promo_bookings_count']}",
        ]
        for line in summary_lines:
            draw_text(line, size=11)

        y -= 3 * mm

        draw_section_title('Популярность номеров')
        if data['charts']['room_type_distribution']:
            for row in data['charts']['room_type_distribution']:
                draw_text(f"{row['name']}: {row['value']}", size=11)
        else:
            draw_text('Нет данных', size=11)

        y -= 3 * mm

        draw_section_title('Промокоды')
        if data['tables']['promotions_report']:
            for row in data['tables']['promotions_report']:
                line = (
                    f"{row['promo_code']} | использований: {row['usages']} | "
                    f"скидки: {money(row['total_discount'])} | доход: {money(row['total_revenue'])}"
                )
                draw_text(line, size=10)
        else:
            draw_text('Нет данных', size=11)

        y -= 3 * mm

        draw_section_title('Топ клиентов')
        if data['tables']['top_clients']:
            for row in data['tables']['top_clients'][:10]:
                line = (
                    f"{row['client']} | броней: {row['bookings_count']} | "
                    f"отмен: {row['cancellations']} | средний чек: {money(row['average_check'])} | "
                    f"потрачено: {money(row['total_spent'])}"
                )
                draw_text(line, size=10)
        else:
            draw_text('Нет данных', size=11)

        y -= 3 * mm

        draw_section_title('Последние бронирования')
        if data['tables']['bookings_table']:
            for row in data['tables']['bookings_table'][:10]:
                line = (
                    f"{row['booking_number']} | {row['client']} | {row['type']} | "
                    f"{money(row['total_price'])} | {row['status']} | "
                    f"{row['start_date']} - {row['end_date']}"
                )
                draw_text(line, size=10)
        else:
            draw_text('Нет данных', size=11)

        p.save()
        buffer.seek(0)

        filename = f"analytics_report_{data['period']['from']}_{data['period']['to']}.pdf"

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response